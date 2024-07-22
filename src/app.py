
from functools import wraps
import io
from io import BytesIO
import os
import pyodbc
from flask import Flask, flash, make_response, redirect, render_template, request, send_file, session, url_for, send_from_directory, jsonify, g
import datetime
from werkzeug.utils import secure_filename
from waitress import serve
from fpdf import FPDF, XPos, YPos
from openpyxl import Workbook
import hashlib
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import configparser
import sys
import re






app = Flask(__name__, template_folder='templates', static_folder='static')

app.config['SECRET_KEY'] = 'chf24'


if getattr(sys, 'frozen', False):
    config_file = os.path.join(os.path.dirname(sys.executable), 'config.ini')
else:
    config_file = os.path.join(os.path.dirname(__file__), 'config.ini')


if not os.path.exists(config_file):
    print(f'Error: El archivo {config_file} no encontrado')
    exit(1)


config = configparser.ConfigParser()
config.read(config_file)


server = config.get('database', 'server')
db = config.get('database', 'db')
user = config.get('database', 'user')
password = config.get('database', 'password')
ruta = config.get('ruta', 'ruta')
###
#Conexion a base de datos-------------------------------
try:
    conexion = pyodbc.connect(
        'DRIVER={ODBC DRIVER 17 FOR SQL server};SERVER=' + server + ';DATABASE=' + db + 
        ';UID=' + user + ';PWD=' + password
    )
    print('Conexión exitosa')
except pyodbc.Error as ex:
    sqlstate = ex.args[0]
    print(sqlstate)
    print ('Error al intentar conectarse')

#------------------------------
def obtener_conexion():

    try:
        conexion = pyodbc.connect(
            'DRIVER={SQL Server};'
            f'SERVER={server};'
            f'DATABASE={db};'
            f'UID={user};'
            f'PWD={password}'
        )
        return conexion
    except pyodbc.Error as ex:
        print(f"Error al intentar conectarse: {ex}")
        return None
############################################################
UPLOAD_FOLDER = 'C:/descarga/'
ALLOWED_EXTENSIONS = {'pdf'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
###

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_email' not in session:
            flash('Por favor, inicie sesión para acceder a esta página.')
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def validate_user(email, password):
    conexion = obtener_conexion()
    cursor = conexion.cursor()
    cursor.execute("SELECT * FROM Usuarios WHERE usuarios_correo = ? AND usuarios_contraseña = ?", (email, password))
    user = cursor.fetchone()
    conexion.close()
    return user

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        correo = request.form['correo']
        contraseña = request.form['contraseña']
        user = validate_user(correo, contraseña)


        if user:
            session['user_email'] = user.usuarios_correo
            session['user_nombre'] = user.usuarios_nombre
            session['user_ap_paterno'] = user.usuarios_ap_paterno
            session['user_ap_materno'] = user.usuarios_ap_materno
            session['login_time'] = datetime.datetime.now().strftime('%Y-%d-%m %H:%M:%S')

            flash(f'Bienvenido {user.usuarios_nombre} {user.usuarios_ap_paterno} {user.usuarios_ap_materno}')
            next_page = request.args.get('next')
            return redirect(next_page or url_for('index', welcome_message=f'Bienvenido {user.usuarios_nombre} {user.usuarios_ap_paterno} {user.usuarios_ap_materno}'))
        else:
            flash('Correo electrónico o contraseña incorrectos.')
    return render_template('login.html', current_date=session.get('current_date'))


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


#################################################



def verificar_duplicado(tabla, columnas_valores):
    conexion = None
    cursor = None
    try:
        conexion = obtener_conexion()
        if not conexion:
            raise Exception("No se pudo conectar a la base de datos")

        cursor = conexion.cursor()
        
        # Crear la parte de la consulta que compara las columnas y valores
        condiciones = " AND ".join([f"{columna} = ?" for columna in columnas_valores.keys()])
        query = f"SELECT COUNT(1) FROM {tabla} WHERE {condiciones}"
        
        cursor.execute(query, tuple(columnas_valores.values()))
        resultado = cursor.fetchone()
        
        return resultado[0] > 0
    except Exception as e:
        print(f"Error al verificar duplicado: {e}")
        return False
    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()



#---------------------------------------------------------
# Index mostrar personas
@app.route('/', methods=['GET', 'POST'])
@login_required
def index():
    welcome_message = request.args.get('welcome_message', '')

    cursor = conexion.cursor()
    try:
        search_term = request.form.get('search_term', '').lower()

        # Consulta para obtener las personas
        persona_query = '''
            SELECT
                persona.id,
                persona.nombres,
                persona.apellidos,
                persona.correo,
                persona.rut,
                persona.dv,
                area.nombre AS area_nombre
            FROM
                persona
            INNER JOIN
                area ON persona.area_id = area.id
            ORDER BY persona.nombres ASC
        '''

        cursor.execute(persona_query)
        personas_rows = cursor.fetchall()

        # Consulta para obtener los equipos asignados
        equipo_query = '''
            SELECT
                asignacion_equipo.persona_id,
                equipo.id AS equipo_id,
                tipoequipo.nombre AS tipo_nombre,
                ISNULL(unidad.nombre_e, '') AS nombre_unidad,
                ISNULL(celular.nombre, '') AS nombre_celular
            FROM
                asignacion_equipo
            INNER JOIN
                equipo ON asignacion_equipo.equipo_id = equipo.id
            LEFT JOIN
                unidad ON equipo.unidad_id = unidad.id
            LEFT JOIN
                celular ON equipo.celular_id = celular.id
            LEFT JOIN
                tipoequipo ON equipo.tipoequipo_id = tipoequipo.id
            WHERE
                asignacion_equipo.fecha_devolucion IS NULL
        '''

        cursor.execute(equipo_query)
        equipos_rows = cursor.fetchall()
        cursor.close()

        # Crear un diccionario para almacenar las personas y sus equipos asignados
        personas = {}
        for row in personas_rows:
            persona_id = row.id
            personas[persona_id] = {
                'id': row.id,
                'nombres': row.nombres,
                'apellidos': row.apellidos,
                'correo': row.correo,
                'rut': row.rut,
                'dv': row.dv,
                'area_nombre': row.area_nombre,
                'equipos_asignados': []
            }

        # Agregar equipos asignados a las personas
        for row in equipos_rows:
            persona_id = row.persona_id
            if persona_id in personas:
                equipo = {
                    'equipo_id': row.equipo_id,
                    'nombre': row.nombre_unidad if row.nombre_unidad else row.nombre_celular,
                    'tipoequipo': row.tipo_nombre
                }
                personas[persona_id]['equipos_asignados'].append(equipo)

        # Filtrar personas y equipos según el término de búsqueda
        if search_term:
            filtered_personas = {}
            for persona_id, persona in personas.items():
                # Buscar en campos de persona
                if (search_term in persona['nombres'].lower() or
                    search_term in persona['apellidos'].lower() or
                    search_term in persona['correo'].lower() or
                    search_term in persona['rut'].lower() or
                    search_term in persona['area_nombre'].lower()):
                    filtered_personas[persona_id] = persona
                else:
                    # Buscar en equipos asignados
                    equipos_coincidencia = [
                        equipo for equipo in persona['equipos_asignados']
                        if search_term in equipo['nombre'].lower() or search_term in equipo['tipoequipo'].lower()
                    ]
                    if equipos_coincidencia:
                        persona['equipos_asignados'] = equipos_coincidencia
                        filtered_personas[persona_id] = persona
        else:
            filtered_personas = personas

        return render_template('index.html', personas=list(filtered_personas.values()), search_term=search_term, welcome_message=welcome_message, login_time=session.get('login_time'))
    except pyodbc.Error as ex:
        print(f'Error al obtener los datos: {ex}')
        flash('Error al cargar la página de asignación', 'error')
        return redirect(url_for('index'))




#--------------------------------------------------------
# CRUD Persona
@app.route('/listar_personas', methods=['GET', 'POST'])
@login_required
def listar_personas():
    search_term = request.args.get('search_term', '')
    try:
        cursor = conexion.cursor()
        query = '''
            SELECT persona.id, persona.nombres, persona.apellidos, persona.correo, area.nombre AS area_nombre
            FROM persona
            INNER JOIN area ON persona.area_id = area.id
            WHERE persona.nombres LIKE ? OR persona.apellidos LIKE ? OR persona.correo LIKE ? OR area.nombre LIKE ?
            ORDER BY persona.nombres ASC
        '''
        search_term_wildcard = '%' + search_term + '%'
        cursor.execute(query, (search_term_wildcard, search_term_wildcard, search_term_wildcard, search_term_wildcard))
        personas = cursor.fetchall()
        cursor.close()
        return render_template('listar_personas.html', personas=personas)
    except pyodbc.Error as ex:
        print(ex)
        return "Error en la consulta de la base de datos"



@app.route('/agregar_persona', methods=['GET', 'POST'])
@login_required
def agregar_persona():
    areas = obtener_areas()

    if request.method == 'POST':
        nombres = request.form['nombres']
        apellidos = request.form['apellidos']
        correo = request.form['correo']
        rut_completo = request.form['rut'].replace(".", "").replace("-", "") if request.form['rut'] else ""
        rut = rut_completo[:-1] if rut_completo else ""
        dv = rut_completo[-1] if rut_completo else ""
        area_id = request.form['area']

        try:
            is_duplicate_comb, is_duplicate_email = verificar_duplicado_persona(nombres, apellidos, correo)
            
            if is_duplicate_comb:
                flash('La combinación de nombre, apellido o correo ya existe en la base de datos', 'a_persona_error')
                return render_template('agregar_persona.html', nombres=nombres, apellidos=apellidos, correo=correo, rut=rut, area_id=area_id, areas=areas)
            
            if is_duplicate_email:
                flash('El correo ya existe en la base de datos', 'a_persona_error')
                return render_template('agregar_persona.html', nombres=nombres, apellidos=apellidos, correo=correo, rut=rut, area_id=area_id, areas=areas)

            cursor = conexion.cursor()
            cursor.execute("INSERT INTO persona (nombres, apellidos, correo, rut, dv, area_id) VALUES (?, ?, ?, ?, ?, ?)",
                           (nombres, apellidos, correo if correo else "", rut, dv, area_id))
            conexion.commit()
            cursor.close()
            return redirect(url_for('listar_personas'))
        except pyodbc.Error as ex:
            print(ex)
            flash('Ocurrió un error al agregar la persona', 'a_persona_error')
            return render_template('agregar_persona.html', nombres=nombres, apellidos=apellidos, correo=correo, rut=rut, area_id=area_id, areas=areas)

    return render_template('agregar_persona.html', areas=areas)

def verificar_duplicado_persona(nombres, apellidos, correo, id=None):
    cursor = conexion.cursor()
    
    query_comb = "SELECT COUNT(*) FROM persona WHERE nombres = ? AND apellidos = ?"
    params_comb = [nombres, apellidos]
    if correo:
        query_comb += " AND correo = ?"
        params_comb.append(correo)
    if id:
        query_comb += " AND id != ?"
        params_comb.append(id)
    
    cursor.execute(query_comb, params_comb)
    count_comb = cursor.fetchone()[0]
    
    count_email = 0
    if correo:
        query_email = "SELECT COUNT(*) FROM persona WHERE correo = ?"
        params_email = [correo]
        if id:
            query_email += " AND id != ?"
            params_email.append(id)
        
        cursor.execute(query_email, params_email)
        count_email = cursor.fetchone()[0]

    cursor.close()
    
    return count_comb > 0, count_email > 0



@app.route('/editar_persona/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_persona(id):
    if request.method == 'POST':
        # Obtener datos del formulario
        nombres = request.form['nombres']
        apellidos = request.form['apellidos']
        correo = request.form['correo']
        rut_completo = request.form['rut'].replace(".", "").replace("-", "")
        rut = rut_completo[:-1] if rut_completo else ""
        dv = rut_completo[-1] if rut_completo else ""
        area_id = request.form['area']

        try:
            cursor = conexion.cursor()
            cursor.execute("UPDATE persona SET nombres=?, apellidos=?, correo=?, rut=?, dv=?, area_id=? WHERE id=?",
                           (nombres, apellidos, correo, rut, dv, area_id, id))
            conexion.commit()
            cursor.close()
            return redirect(url_for('listar_personas'))
        except pyodbc.Error as ex:
            print(ex)
            flash('Ocurrió un error al actualizar los datos de la persona', 'error')
            return redirect(url_for('editar_persona', id=id))

    try:
        cursor = conexion.cursor()
        cursor.execute("SELECT * FROM persona WHERE id=?", (id,))
        persona = cursor.fetchone()
        areas = obtener_areas()
        cursor.close()
        return render_template('editar_persona.html', persona=persona, areas=areas)
    except pyodbc.Error as ex:
        print(ex)
        flash('Ocurrió un error al obtener los datos de la persona', 'error')
        return redirect(url_for('listar_personas'))






@app.route('/eliminar_persona/<int:id>', methods=['POST'])
@login_required
def eliminar_persona(id):
    try:
        cursor = conexion.cursor()
        cursor.execute("DELETE FROM persona WHERE id = ?", (id,))
        conexion.commit()
        cursor.close()
    except pyodbc.Error as ex:
        print(ex)
    return redirect(url_for('listar_personas'))

#--------------------------------------------------------
#Barra de busqueda
@app.route('/buscar_personas')
def buscar_personas():
    q = request.args.get('q', '')
    try:
        cursor = conexion.cursor()
        cursor.execute('''
            SELECT * FROM Persona
            WHERE nombres LIKE ? OR correo LIKE ? OR area LIKE ?
        ''', ('%' + q + '%', '%' + q + '%', '%' + q + '%'))
        personas = cursor.fetchall()
        cursor.close()
        return render_template('index.html', personas=personas)
    except pyodbc.Error as ex:
        print(ex)
        return 'Error al buscar personas'
#----------------------------------------
# CRUD de equipos:
@app.route('/mostrar_equipos', methods=['GET'])
def mostrar_equipos():
    search_term = request.args.get('search_term', '').strip()

    try:
        cursor = conexion.cursor()
        query = """
        SELECT
            equipo.id,
            estadoequipo.nombre AS estado_nombre,
            tipoequipo.nombre AS tipo_nombre,
            COALESCE(unidad.nombre_e, celular.nombre) AS nombre_equipo,
            COALESCE(unidad.serial, celular.serial) AS sn_equipo,
            COALESCE(unidad.modelo, celular.modelo) AS modelo_equipo
        FROM equipo
        LEFT JOIN tipoequipo ON equipo.tipoequipo_id = tipoequipo.id
        LEFT JOIN unidad ON equipo.unidad_id = unidad.id
        LEFT JOIN celular ON equipo.celular_id = celular.id
        LEFT JOIN estadoequipo ON equipo.estadoequipo_id = estadoequipo.id
        WHERE
            estadoequipo.nombre LIKE ? OR
            tipoequipo.nombre LIKE ? OR
            unidad.nombre_e LIKE ? OR
            celular.nombre LIKE ? OR
            unidad.serial LIKE ? OR
            celular.serial LIKE ? OR
            unidad.modelo LIKE ? OR
            celular.modelo LIKE ?
        """
        search_term_wildcard = '%' + search_term + '%'
        cursor.execute(query, (search_term_wildcard, search_term_wildcard, search_term_wildcard, search_term_wildcard,
                               search_term_wildcard, search_term_wildcard, search_term_wildcard, search_term_wildcard))
        equipos = cursor.fetchall()
        cursor.close()
        return render_template('mostrar_equipos.html', equipos=equipos)
    except pyodbc.Error as ex:
        print(f'Error al obtener los equipos: {ex}')
        flash('Error al obtener los equipos', 'error')
        return redirect(url_for('index'))

# Ruta para buscar equipos dinámicamente
@app.route('/buscar_equipos', methods=['GET'])
def buscar_equipos():
    search_term = request.args.get('search_term', '').strip()

    try:
        cursor = conexion.cursor()
        query = """
        SELECT
            equipo.id,
            estadoequipo.nombre AS estado_nombre,
            tipoequipo.nombre AS tipo_nombre,
            COALESCE(unidad.nombre_e, celular.nombre) AS nombre_equipo,
            COALESCE(unidad.serial, celular.serial) AS sn_equipo,
            COALESCE(unidad.modelo, celular.modelo) AS modelo_equipo
        FROM equipo
        LEFT JOIN tipoequipo ON equipo.tipoequipo_id = tipoequipo.id
        LEFT JOIN unidad ON equipo.unidad_id = unidad.id
        LEFT JOIN celular ON equipo.celular_id = celular.id
        LEFT JOIN estadoequipo ON equipo.estadoequipo_id = estadoequipo.id
        WHERE
            estadoequipo.nombre LIKE ? OR
            tipoequipo.nombre LIKE ? OR
            unidad.nombre_e LIKE ? OR
            celular.nombre LIKE ? OR
            unidad.serial LIKE ? OR
            celular.serial LIKE ? OR
            unidad.modelo LIKE ? OR
            celular.modelo LIKE ?
        """
        search_term_wildcard = '%' + search_term + '%'
        cursor.execute(query, (search_term_wildcard, search_term_wildcard, search_term_wildcard, search_term_wildcard,
                               search_term_wildcard, search_term_wildcard, search_term_wildcard, search_term_wildcard))
        equipos = cursor.fetchall()
        cursor.close()

        # Renderizar solo el cuerpo de la tabla de equipos para la respuesta AJAX
        return render_template('equipos_table_body.html', equipos=equipos)
    except pyodbc.Error as ex:
        print(f'Error al buscar equipos: {ex}')
        return jsonify({'error': 'Error al buscar equipos'})



#------------------------------------------------------------
#Obtener tipos de equipo
def obtener_tipos_equipo():
    cursor = conexion.cursor()
    cursor.execute('SELECT id, nombre FROM tipoequipo ORDER BY nombre ASC ')
    tipos_equipo = cursor.fetchall()
    return tipos_equipo
#------------------------------------------------------------
#Obtener estados
def obtener_estado():
    cursor = conexion.cursor()
    cursor.execute('SELECT id, nombre FROM estadoequipo')
    estado_equipo = cursor.fetchall()
    return estado_equipo
#------------------------------------------------------------


@app.route('/agregar_equipo', methods=['GET', 'POST'])
def agregar_equipo():
    tipos_equipos = obtener_tipos_equipo()

    if request.method == 'POST':
        tipo_equipo = request.form.get('tipo_equipo')
        if '_' in tipo_equipo:
            tipo_equipo_id, tipo_equipo_nombre = tipo_equipo.split('_', 1)
            tipo_equipo_nombre = tipo_equipo_nombre.replace('_', ' ')  # Reemplazar guiones bajos con espacios
        else:
            # Manejar caso cuando no hay guiones bajos
            tipo_equipo_id = tipo_equipo
            tipo_equipo_nombre = tipo_equipo
        observaciones = request.form.get('observaciones', '')

        nombre = request.form.get('nombre', '').upper()
        marca = request.form.get('marca', '').upper()
        modelo = request.form.get('modelo', '')
        serial = request.form.get('serial', '').upper()

        # Limpiar el nombre
        nombre_limpio = re.sub(r'[^A-Za-z0-9]', '', nombre)

        try:
            # Verificar si el nombre del equipo ya existe en las tablas 'Unidad' o 'Celular'
            if nombre_limpio:
                if verificar_duplicado('Unidad', {'nombre_e': nombre_limpio}) or verificar_duplicado('Celular', {'nombre': nombre_limpio}):
                    flash('El nombre del equipo ya existe en la base de datos', 'agregar_equipo_error')
                    return render_template('agregar_equipo.html', tipos_equipos=tipos_equipos, form_data=request.form)

            if tipo_equipo_nombre in ['notebook', 'pc', 'mac', 'macbook']:
                ram = request.form.get('ram', '')
                procesador = request.form.get('procesador', '')
                almc = request.form.get('almc', '')
                perif = request.form.get('perif', '')
                numsello = request.form.get('numsello', '')
                numproducto = request.form.get('numproducto', '').upper()
                insertar_equipo(nombre_limpio, marca, modelo, ram, procesador, almc, perif, numsello, serial, numproducto, None, None, observaciones, tipo_equipo_id)

            elif tipo_equipo_nombre == 'monitor':
                numproducto = request.form.get('numproducto', '').upper()
                insertar_equipo(nombre_limpio, marca, modelo, None, None, None, None, None, serial, numproducto, None, None, observaciones, tipo_equipo_id)

            elif tipo_equipo_nombre == 'impresora':
                tipoimpresion = request.form.get('tipoimpresion', '')
                perif = request.form.get('perif', '')
                insertar_equipo(nombre_limpio, marca, modelo, None, None, None, perif, None, serial, None, tipoimpresion, None, observaciones, tipo_equipo_id)

            elif tipo_equipo_nombre == 'celular':
                imei1 = request.form.get('imei1', '')
                imei2 = request.form.get('imei2', '')
                insertar_celular(nombre_limpio, marca, modelo, imei1, imei2, serial, None, observaciones, tipo_equipo_id)

            elif tipo_equipo_nombre == 'accesorios':
                cantidad = request.form.get('cantidad', '')
                insertar_equipo(nombre_limpio, marca, modelo, None, None, None, None, None, serial, None, None, cantidad, observaciones, tipo_equipo_id)

            elif tipo_equipo_nombre == 'simcard':
                imei = request.form.get('imei', '')
                ntelefono = request.form.get('ntelefono', '')
                insertar_celular(imei, None, None, imei, None, None, ntelefono, observaciones, tipo_equipo_id)
            
            else:
                insertar_equipo(nombre_limpio, marca, modelo, None, None, None, None, None, serial, None, None, None, observaciones, tipo_equipo_id)

            return redirect(url_for('mostrar_equipos'))

        except Exception as e:
            flash(f'Ocurrió un error al agregar el equipo: {str(e)}','agregar_equipo_error')
            print(e)
            return render_template('agregar_equipo.html', tipos_equipos=tipos_equipos, form_data=request.form)

    return render_template('agregar_equipo.html', tipos_equipos=tipos_equipos)






def insertar_equipo(nombre, marca, modelo, ram, procesador, almc, perif, numsello, serial, numproducto, tipoimpresion, cantidad, observaciones, tipo_equipo_id):
    try:
        cursor = conexion.cursor()
        cursor.execute("INSERT INTO unidad (nombre_e, marca, modelo, ram, procesador, almc, perif, numsello, serial, numproducto, tipoimpresion, cantidad) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                       (nombre, marca, modelo, ram, procesador, almc, perif, numsello, serial, numproducto, tipoimpresion, cantidad))
        conexion.commit()
        cursor.execute("SELECT @@IDENTITY")
        unidad_id = cursor.fetchone()[0]

        cursor.execute("INSERT INTO equipo (estadoequipo_id, unidad_id, fcreacion, tipoequipo_id, observaciones) VALUES (?, ?, ?, ?, ?)",
                       (1, unidad_id, datetime.datetime.now(), tipo_equipo_id, observaciones))
        conexion.commit()
        cursor.close()
    except Exception as e:
        print("Error al insertar equipo:", e)

def insertar_celular(nombre, marca, modelo, imei1, imei2, serial, ntelefono, observaciones, tipo_equipo_id):
    cursor = conexion.cursor()
    cursor.execute("INSERT INTO celular (nombre, marca, modelo, imei1, imei2, serial, ntelefono) VALUES (?, ?, ?, ?, ?, ?, ?)",
                   (nombre, marca, modelo, imei1, imei2, serial, ntelefono))
    conexion.commit()
    cursor.execute("SELECT @@IDENTITY")
    celular_id = cursor.fetchone()[0]

    cursor.execute("INSERT INTO equipo (estadoequipo_id, celular_id, fcreacion, tipoequipo_id, observaciones) VALUES (?, ?, ?, ?, ?)",
                   (1, celular_id, datetime.datetime.now(), tipo_equipo_id, observaciones))
    conexion.commit()
    cursor.close()

#---------------------------------------------------
#Editar equipo:
@app.route('/editar_equipo/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_equipo(id):
    detalle_equipo = None

    if request.method == 'GET':
        try:
            cursor = conexion.cursor()
            cursor.execute("SELECT * FROM equipo WHERE id = ?", (id,))
            equipo = cursor.fetchone()

            if equipo:
                equipo_dict = dict(zip((column[0] for column in cursor.description), equipo))
                detalle_equipo = None
                if equipo_dict['unidad_id'] is not None:
                    cursor.execute("SELECT equipo.*, unidad.* FROM equipo INNER JOIN unidad ON equipo.unidad_id = unidad.id WHERE equipo.id = ?", (id,))
                    detalle_equipo = cursor.fetchone()
                    tipo_equipo = 'unidad'
                elif equipo_dict['celular_id'] is not None:
                    cursor.execute("SELECT equipo.*, celular.* FROM equipo INNER JOIN celular ON equipo.celular_id = celular.id WHERE equipo.id = ?", (id,))
                    detalle_equipo = cursor.fetchone()
                    tipo_equipo = 'celular'

                estados = obtener_estados_equipo()

                return render_template('editar_equipo.html', equipo=equipo_dict, detalle_equipo=detalle_equipo,
                                       estados=estados, tipo_equipo=tipo_equipo)
            else:
                flash('No se encontró ningún equipo con el ID proporcionado', 'error')
                return redirect(url_for('mostrar_equipos'))

        except pyodbc.Error as ex:
            print(ex)
            flash('Ocurrió un error al cargar la información del equipo', 'error')
            return redirect(url_for('mostrar_equipos'))

        finally:
            cursor.close()

    elif request.method == 'POST':
        try:
            cursor = conexion.cursor()
            cursor.execute("SELECT * FROM equipo WHERE id = ?", (id,))
            equipo = cursor.fetchone()

            if equipo:
                equipo_dict = dict(zip((column[0] for column in cursor.description), equipo))

                nombre = request.form.get('nombre', equipo_dict.get('nombre', '')).upper()
                marca = request.form.get('marca', equipo_dict.get('marca', ''))
                modelo = request.form.get('modelo', equipo_dict.get('modelo', ''))
                serial = request.form.get('serial', equipo_dict.get('serial', '')).upper()
                observaciones = request.form['observaciones']

                # Limpiar el nombre
                nombre_limpio = re.sub(r'[^A-Za-z0-9]', '', nombre)

                if 'unidad_id' in equipo_dict and equipo_dict['unidad_id'] is not None:
                    cursor.execute(
                        "UPDATE unidad SET nombre_e = ?, marca = ?, modelo = ?, serial = ?, ram = ?, procesador = ?, almc = ?, perif = ?, numsello = ?, numproducto = ?, tipoimpresion = ?, cantidad = ? WHERE id = ?",
                        (nombre_limpio, marca, modelo, serial, request.form.get('ram', ''), request.form.get('procesador', ''),
                         request.form.get('almc', ''), request.form.get('perif', ''), request.form.get('numsello', ''),
                         request.form.get('numproducto', ''), request.form.get('tipoimpresion', ''),
                         request.form.get('cantidad', ''), equipo_dict['unidad_id']))
                elif 'celular_id' in equipo_dict and equipo_dict['celular_id'] is not None:
                    cursor.execute(
                        "UPDATE celular SET nombre = ?, marca = ?, modelo = ?, serial = ?, imei1 = ?, imei2 = ?, ntelefono = ? WHERE id = ?",
                        (nombre_limpio, marca, modelo, serial, request.form.get('imei1', ''), request.form.get('imei2', ''),
                         request.form.get('ntelefono', ''), equipo_dict['celular_id']))

                cursor.execute("UPDATE equipo SET observaciones = ? WHERE id = ?",
                               (observaciones, id))

                conexion.commit()
                cursor.close()

                return redirect(url_for('mostrar_equipos'))
            else:
                return redirect(url_for('mostrar_equipos'))

        except pyodbc.Error as ex:
            print(ex)
            flash('Ocurrió un error al intentar actualizar la información del equipo', 'error')
            return redirect(url_for('mostrar_equipos'))

    else:
        flash('Método HTTP no permitido', 'error')
        return redirect(url_for('mostrar_equipos'))



#-------------------------------------------------------
#Funcion Eliminar equipo
@app.route('/eliminar_equipo/<int:id>', methods=['POST'])
@login_required
def eliminar_equipo(id):
    conexion = None
    cursor = None
    try:
        conexion = obtener_conexion()
        if not conexion:
            raise Exception("No se pudo conectar a la base de datos")
        
        cursor = conexion.cursor()
        print(f"Conexión obtenida, eliminando equipo con ID: {id}")

        # Primero, obtenemos los IDs relacionados en unidad y celular
        cursor.execute("SELECT unidad_id, celular_id FROM equipo WHERE id = ?", (id,))
        equipo = cursor.fetchone()
        
        if equipo:
            unidad_id, celular_id = equipo

            # Eliminamos los registros relacionados en unidad y celular si existen
            if unidad_id:
                cursor.execute("DELETE FROM unidad WHERE id = ?", (unidad_id,))
            if celular_id:
                cursor.execute("DELETE FROM celular WHERE id = ?", (celular_id,))

            # Luego, eliminamos el registro de la tabla equipo
            cursor.execute("DELETE FROM equipo WHERE id = ?", (id,))
            conexion.commit()

            print(f"Equipo con ID: {id} eliminado correctamente")
            flash('El equipo ha sido eliminado correctamente', 'success')
        else:
            print(f"No se encontró el equipo con ID: {id}")
            flash('No se encontró el equipo a eliminar', 'error')

    except pyodbc.Error as ex:
        if conexion:
            conexion.rollback()  # Revertir los cambios en caso de error
        print(f"Error al eliminar el equipo: {ex}")
        flash('Ocurrió un error al eliminar el equipo', 'error')
    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()
    
    return redirect(url_for('mostrar_equipos'))


#--------------------------------------------------
# Obtener estados
def obtener_estados_equipo():
    try:
        cursor = conexion.cursor()
        cursor.execute("SELECT id, nombre FROM estadoequipo")
        estequipos = cursor.fetchall()
        cursor.close()
        return estequipos
    except pyodbc.Error as ex:
        print(ex)
        return None
#---------------------------------------------------
#CRUD Area

@app.route('/listar_areas', methods=['GET', 'POST'])
@login_required
def listar_areas():
    if request.method == 'GET':
        try:
            cursor = conexion.cursor()
            cursor.execute('SELECT * FROM area')
            areas = cursor.fetchall()
            cursor.close()
            return render_template('listar_areas.html', areas=areas)
        except pyodbc.Error as ex:
            print(ex)
            flash('Ocurrió un error al listar las áreas', 'area_error')
            return redirect(url_for('index'))
    elif request.method == 'POST':
        id_area = request.form.get('id_area')
        nombre = request.form.get('nombre')
        try:
            cursor = conexion.cursor()
            cursor.execute("UPDATE area SET nombre = ? WHERE id = ?", (nombre, id_area))
            conexion.commit()
            cursor.close()
            flash('Área actualizada correctamente', 'area_success')
        except pyodbc.Error as ex:
            print(ex)
            flash('Error al actualizar el área', 'area_error')
        return redirect(url_for('listar_areas'))

@app.route('/crear_area', methods=['GET', 'POST'])
@login_required
def crear_area():
    if request.method == 'POST':
        nombre = request.form['nombre']
        try:
            cursor = conexion.cursor()
            cursor.execute("INSERT INTO area (nombre) VALUES (?)", (nombre,))
            conexion.commit()
            cursor.close()
            flash('Área creada correctamente', 'area_success')
            return redirect(url_for('listar_areas'))
        except pyodbc.Error as ex:
            print(ex)
            flash('Error al crear el área', 'area_error')
            return redirect(url_for('listar_areas'))
    return render_template('listar_areas.html')


@app.route('/eliminar_area/<int:id>', methods=['POST'])
@login_required
def eliminar_area(id):
    if request.method == 'POST':
        try:
            cursor = conexion.cursor()
            cursor.execute("DELETE FROM area WHERE id = ?", (id,))
            conexion.commit()
            cursor.close()
            flash('Área eliminada correctamente', 'area_success')
            return redirect(url_for('listar_areas'))
        except pyodbc.Error as ex:
            print(ex)
            flash('Error al eliminar el área', 'area_error')
            return redirect(url_for('listar_areas'))
    return 'Método no permitido'
#-----------------------------------------------------------------
#CRUD TIPOS
@app.route('/crud_tipos_equipos', methods=['GET', 'POST'])
@login_required
def crud_tipos_equipos():
    if request.method == 'POST':
        if 'agregar' in request.form:
            # Agregar un nuevo tipo de equipo
            nombre = request.form['nombre']
            if nombre:
                try:
                    cursor = conexion.cursor()
                    cursor.execute("INSERT INTO tipoequipo (nombre) VALUES (?)", (nombre,))
                    conexion.commit()
                    cursor.close()
                    flash('El tipo de equipo ha sido agregado correctamente', 'tiposE_success')
                except pyodbc.Error as ex:
                    print(ex)
                    flash('Ocurrió un error al agregar el tipo de equipo', 'tiposE_error')
            else:
                flash('Por favor, ingrese un nombre para el tipo de equipo', 'tiposE_error')
        elif 'eliminar' in request.form:
            tipo_equipo_id = request.form.get('eliminar')
            try:
                cursor = conexion.cursor()
                cursor.execute("DELETE FROM tipoequipo WHERE id = ?", (tipo_equipo_id,))
                conexion.commit()
                cursor.close()
                flash('El tipo de equipo ha sido eliminado correctamente', 'tiposE_success')
            except pyodbc.Error as ex:
                print(ex)
                flash('Ocurrió un error al eliminar el tipo de equipo', 'tiposE_error')
        elif 'editar' in request.form:
            tipo_equipo_id = request.form['editar']
            nuevo_nombre = request.form['nombre_editado']
            
            if nuevo_nombre:
                try:
                    cursor = conexion.cursor()
                    cursor.execute("UPDATE tipoequipo SET nombre = ? WHERE id = ?", (nuevo_nombre, tipo_equipo_id))
                    conexion.commit()
                    cursor.close()
                    flash('El tipo de equipo ha sido actualizado correctamente', 'tiposE_success')
                except pyodbc.Error as ex:
                    print(ex)
                    flash('Ocurrió un error al actualizar el tipo de equipo', 'tiposE_error')
            else:
                flash('Por favor, ingrese un nuevo nombre para el tipo de equipo', 'tiposE_error')

    tipos_equipos = obtener_tipos_equipo()

    return render_template('crud_tipos_equipos.html', tipos_equipos=tipos_equipos)

#--------------------------------------------------------------------------------
# Obtener areas
def obtener_areas():
    try:
        cursor = conexion.cursor()
        cursor.execute("SELECT id, nombre FROM area ORDER BY nombre ASC")
        areas = cursor.fetchall()
        cursor.close()
        return areas
    except pyodbc.Error as ex:
        print(ex)
        return None

#----------------------------------------------------------
#Asignar equipo

@app.route('/asignar_equipo', methods=['GET', 'POST'])
@login_required
def asignar_equipo():
    cursor = conexion.cursor()

    if request.method == 'POST':
        equipo_ids = request.form.getlist('equipo_id[]')
        persona_id = request.form['persona_id']
        fecha_asignacion = datetime.datetime.now().strftime('%Y-%d-%m %H:%M:%S')
        observaciones = request.form['observaciones']
        archivo_pdf = request.files['archivo_pdf']

        # Verificación de duplicados en el backend
        if len(equipo_ids) != len(set(equipo_ids)):
            flash('No se pueden asignar equipos duplicados.', 'asignar_equipo_error')
            return redirect(url_for('asignar_equipo', persona_id=persona_id))

        # Verificar que el archivo es un PDF
        if archivo_pdf and archivo_pdf.filename.endswith('.pdf'):
            try:
                archivo_pdf.seek(0)  # Asegurarse de estar al inicio del archivo antes de leerlo

                for equipo_id in equipo_ids:
                    if not equipo_id:  # Verificar que equipo_id no esté vacío
                        flash('Todos los campos de equipo deben estar seleccionados.', 'asignar_equipo_error')
                        return redirect(url_for('asignar_equipo', persona_id=persona_id))

                    # Verificar que el equipo_id existe en la tabla equipo y obtener el nombre del equipo
                    cursor.execute("""
                        SELECT 
                            COALESCE(unidad.nombre_e, celular.nombre) AS nombre_equipo
                        FROM equipo
                        LEFT JOIN unidad ON equipo.unidad_id = unidad.id
                        LEFT JOIN celular ON equipo.celular_id = celular.id
                        WHERE equipo.id = ?
                    """, (equipo_id,))
                    equipo_info = cursor.fetchone()
                    if equipo_info is None:
                        flash(f'El equipo con ID {equipo_id} no existe.', 'asignar_equipo_error')
                        return redirect(url_for('asignar_equipo', persona_id=persona_id))
                    
                    

                    # Obtener la información de la persona
                    cursor.execute("SELECT nombres, apellidos FROM persona WHERE id = ?", (persona_id,))
                    persona_info = cursor.fetchone()
                    if persona_info is None:
                        flash(f'La persona con ID {persona_id} no existe.', 'asignar_equipo_error')
                        return redirect(url_for('asignar_equipo', persona_id=persona_id))
                    
                    nombres_persona = persona_info[0].replace(' ', '_')
                    apellidos_persona = persona_info[1].replace(' ', '_')

                    # Crear el nombre del archivo
                    filename = f"ActaEntrega_{nombres_persona}_{apellidos_persona}_{datetime.datetime.now().strftime('%d%m%Y')}.pdf"
                    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

                    # Guardar el archivo PDF con el nombre específico
                    archivo_pdf.seek(0)  # Volver al inicio del archivo
                    archivo_pdf.save(filepath)

                    # Inserción en asignacion_equipo
                    cursor.execute("""
                        INSERT INTO asignacion_equipo (equipo_id, persona_id, fecha_asignacion, observaciones, archivo_pdf)
                        VALUES (?, ?, ?, ?, ?)
                    """, (equipo_id, persona_id, fecha_asignacion, observaciones, filename))
                    conexion.commit()

                    # Actualización del estado del equipo
                    cursor.execute("UPDATE equipo SET estadoequipo_id = 2 WHERE id = ?", (equipo_id,))
                    conexion.commit()

                flash('Equipos asignados correctamente', 'asignar_equipo_success')
                return redirect(url_for('index'))
            except pyodbc.Error as ex:
                print(f'Error al asignar el equipo: {ex}')
                flash('Error al asignar los equipos', 'asignar_equipo_error')
                return redirect(url_for('asignar_equipo'))
            finally:
                cursor.close()
        else:
            flash('Por favor verifique el archivo PDF', 'asignar_equipo_error')
            return redirect(url_for('asignar_equipo', persona_id=persona_id))

    else:
        try:
            cursor.execute("""
            SELECT
                equipo.id AS equipo_id,
                tipoequipo.nombre AS tipo_nombre,
                COALESCE(unidad.nombre_e, celular.nombre) AS nombre_equipo
            FROM equipo
            LEFT JOIN tipoequipo ON equipo.tipoequipo_id = tipoequipo.id
            LEFT JOIN unidad ON equipo.unidad_id = unidad.id
            LEFT JOIN celular ON equipo.celular_id = celular.id
            WHERE equipo.estadoequipo_id = 1
        """)
            equipos = cursor.fetchall()

            # Obtener el ID de la persona de la URL
            persona_id = request.args.get('persona_id')

            cursor.execute("SELECT nombres, apellidos FROM persona WHERE id = ?", (persona_id,))
            persona_info = cursor.fetchone()
            nombres_persona = persona_info[0] if persona_info else ""
            apellidos_persona = persona_info[1] if persona_info else ""

            return render_template('asignar_equipo.html', persona_id=persona_id, nombres=nombres_persona, apellidos=apellidos_persona, equipos=equipos)
        except pyodbc.Error as ex:
            print(f'Error al obtener los datos: {ex}')
            flash('Error al cargar la página de asignación', 'error')
            return redirect(url_for('index'))
        finally:
            cursor.close()











#--------------------------------------------------------------------------
# Devolucion:
@app.route('/devolver_equipo', methods=['GET', 'POST'])
@app.route('/devolver_equipo/<int:equipo_id>', methods=['GET', 'POST'])
@login_required
def devolver_equipo(equipo_id=None):
    cursor = conexion.cursor()

    if request.method == 'POST':
        if 'archivo_pdf' not in request.files:
            flash('No se ha seleccionado ningún archivo', 'devolver_error')
            return redirect(request.url)

        file = request.files['archivo_pdf']

        if file.filename == '':
            flash('No se ha seleccionado ningún archivo', 'devolver_error')
            return redirect(request.url)

        if file and allowed_file(file.filename):
            filename = f"devolucion_{datetime.datetime.now().strftime('%Y%m%d')}.pdf"
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))

            fecha_devolucion = datetime.datetime.now().strftime('%Y-%d-%m %H:%M:%S')
            observaciones = request.form['observaciones']
            equipos_a_devolver = request.form.getlist('equipos')

            try:
                if equipo_id:
                    equipos_a_devolver = [equipo_id]
                
                for equipo_id in equipos_a_devolver:
                    cursor.execute("""
                        UPDATE asignacion_equipo
                        SET fecha_devolucion = ?, observaciones = ?, archivo_pdf_devolucion = ?
                        WHERE equipo_id = ? AND fecha_devolucion IS NULL
                    """, (fecha_devolucion, observaciones, filename, equipo_id))
                    conexion.commit()

                    cursor.execute("UPDATE equipo SET estadoequipo_id = 1 WHERE id = ?", (equipo_id,))
                    conexion.commit()

                if equipo_id:
                    persona_id = obtener_persona_id_asociada(equipo_id)
                else:
                    persona_id = request.form['persona_id']

                flash('Equipos devueltos correctamente', 'devolver_success')
                return redirect(url_for('detalle_persona', persona_id=persona_id))
            except pyodbc.Error as ex:
                print(f'Error al devolver los equipos: {ex}')
                flash('Error al devolver los equipos', 'devolver_error')
                return redirect(url_for('devolver_equipo', equipo_id=equipo_id))
            finally:
                cursor.close()

    else:
        if equipo_id:
            try:
                cursor.execute("""
                    SELECT
                        ae.id,
                        e.id AS equipo_id,
                        te.nombre AS tipo_nombre,
                        COALESCE(u.nombre_e, c.nombre) AS nombre_equipo,
                        p.nombres,
                        ae.fecha_asignacion,
                        ae.fecha_devolucion,
                        ae.observaciones,
                        ae.archivo_pdf_devolucion,
                        ae.persona_id
                    FROM asignacion_equipo AS ae
                    JOIN equipo AS e ON ae.equipo_id = e.id
                    JOIN tipoequipo AS te ON e.tipoequipo_id = te.id
                    LEFT JOIN unidad AS u ON e.unidad_id = u.id
                    LEFT JOIN celular AS c ON e.celular_id = c.id
                    JOIN persona AS p ON ae.persona_id = p.id
                    WHERE ae.fecha_devolucion IS NULL AND ae.equipo_id = ?
                """, (equipo_id,))
                asignacion = cursor.fetchone()

                if not asignacion:
                    flash('No se encontró una asignación activa para este equipo', 'error')
                    return redirect(url_for('detalle_equipo', equipo_id=equipo_id))

                persona_id = asignacion[-1]

                return render_template('devolver_equipo.html', asignacion=asignacion, equipo_id=equipo_id, persona_id=persona_id)
            except pyodbc.Error as ex:
                print(f'Error al obtener los datos: {ex}')
                flash('Error al cargar la página de devolución', 'error')
                return redirect(url_for('detalle_equipo', equipo_id=equipo_id))
            finally:
                cursor.close()
        else:
            persona_id = request.args.get('persona_id')
            try:
                cursor.execute("""
                    SELECT
                        ae.id,
                        e.id AS equipo_id,
                        te.nombre AS tipo_nombre,
                        COALESCE(u.nombre_e, c.nombre) AS nombre_equipo,
                        p.nombres,
                        ae.fecha_asignacion,
                        ae.fecha_devolucion,
                        ae.observaciones,
                        ae.archivo_pdf_devolucion,
                        ae.persona_id
                    FROM asignacion_equipo AS ae
                    JOIN equipo AS e ON ae.equipo_id = e.id
                    JOIN tipoequipo AS te ON e.tipoequipo_id = te.id
                    LEFT JOIN unidad AS u ON e.unidad_id = u.id
                    LEFT JOIN celular AS c ON e.celular_id = c.id
                    JOIN persona AS p ON ae.persona_id = p.id
                    WHERE ae.fecha_devolucion IS NULL AND ae.persona_id = ?
                """, (persona_id,))
                equipos = cursor.fetchall()

                if not equipos:
                    flash('No se encontraron asignaciones activas para esta persona', 'error')
                    return redirect(url_for('detalle_persona', persona_id=persona_id))

                return render_template('devolver_equipo.html', equipos=equipos, persona_id=persona_id)
            except pyodbc.Error as ex:
                print(f'Error al obtener los datos: {ex}')
                flash('Error al cargar la página de devolución', 'error')
                return redirect(url_for('detalle_persona', persona_id=persona_id))
            finally:
                cursor.close()



#---
def obtener_persona_id_asociada(equipo_id):
    cursor = conexion.cursor()
    try:
        cursor.execute("SELECT persona_id FROM asignacion_equipo WHERE equipo_id = ?", (equipo_id,))
        persona_id = cursor.fetchone()[0]
        return persona_id
    except pyodbc.Error as ex:
        print(f'Error al obtener el ID de la persona asociada al equipo: {ex}')
        # Manejo del error: puedes retornar un valor predeterminado, lanzar una excepción, o retornar None
        return None

#---------------------------------------------------------------------------
#Detalle persona
@app.route('/detalle_persona/<int:persona_id>')
@login_required
def detalle_persona(persona_id):
    try:
        cursor = conexion.cursor()
        
        # Obtener los detalles de la persona
        cursor.execute("""
            SELECT
                id,
                nombres,
                apellidos,
                correo,
                rut,
                dv
            FROM persona
            WHERE id = ?
        """, (persona_id,))
        persona = cursor.fetchone()

        if not persona:
            flash('No se encontró la persona', 'error')
            return redirect(url_for('index'))

        # Obtener el historial completo de asignaciones y devoluciones de la persona
        cursor.execute("""
            SELECT
                asignacion_equipo.fecha_asignacion,
                asignacion_equipo.fecha_devolucion,
                asignacion_equipo.observaciones,
                asignacion_equipo.archivo_pdf,
                asignacion_equipo.archivo_pdf_devolucion,
                equipo.id AS equipo_id,
                tipoequipo.nombre AS tipo_equipo,
                COALESCE(unidad.nombre_e, celular.nombre) AS nombre_equipo
            FROM asignacion_equipo
            JOIN equipo ON asignacion_equipo.equipo_id = equipo.id
            LEFT JOIN unidad ON equipo.unidad_id = unidad.id
            LEFT JOIN celular ON equipo.celular_id = celular.id
            LEFT JOIN tipoequipo ON equipo.tipoequipo_id = tipoequipo.id
            WHERE asignacion_equipo.persona_id = ?
            ORDER BY asignacion_equipo.fecha_asignacion DESC
        """, (persona_id,))
        historial_completo = cursor.fetchall()

        # Obtener la lista de equipos mantenidos por la persona (solo los asignados actualmente)
        cursor.execute("""
            SELECT
                equipo.id,
                tipoequipo.nombre AS tipo_equipo,
                COALESCE(unidad.nombre_e, celular.nombre) AS nombre_equipo,
                COALESCE(unidad.modelo, celular.modelo) AS modelo_equipo
            FROM equipo
            JOIN asignacion_equipo ON equipo.id = asignacion_equipo.equipo_id
            LEFT JOIN unidad ON equipo.unidad_id = unidad.id
            LEFT JOIN celular ON equipo.celular_id = celular.id
            LEFT JOIN tipoequipo ON equipo.tipoequipo_id = tipoequipo.id
            WHERE asignacion_equipo.persona_id = ?
            AND asignacion_equipo.fecha_devolucion IS NULL
        """, (persona_id,))
        equipos_mantenidos = cursor.fetchall()

        cursor.close()

        return render_template('detalle_persona.html', persona=persona, historial_completo=historial_completo, equipos_mantenidos=equipos_mantenidos)
    
    except pyodbc.Error as ex:
        print(f'Error al obtener los detalles de la persona: {ex}')
        flash('Error al obtener los detalles de la persona', 'error')
        return redirect(url_for('index'))





#----------------------------------------------------
# Detalle equipo

@app.route('/detalle_equipo/<int:equipo_id>')
@login_required
def detalle_equipo(equipo_id):
    try:
        cursor = conexion.cursor()

        cursor.execute("""
            SELECT
                equipo.id,
                estadoequipo.nombre AS estado_nombre,
                equipo.fcreacion AS fecha_creacion,
                equipo.observaciones,
                equipo.unidad_id,
                unidad.nombre_e,
                unidad.marca,
                unidad.modelo,
                unidad.ram,
                unidad.procesador,
                unidad.almc,
                unidad.perif,
                unidad.numsello,
                unidad.serial,
                unidad.numproducto,
                unidad.tipoimpresion,
                unidad.cantidad,
                equipo.celular_id,
                celular.nombre AS nombre_celular,
                celular.marca AS marca_celular,
                celular.modelo AS modelo_celular,
                celular.serial AS serial_celular,
                celular.imei1,
                celular.imei2,
                celular.ntelefono
            FROM equipo
            LEFT JOIN estadoequipo ON equipo.estadoequipo_id = estadoequipo.id
            LEFT JOIN unidad ON equipo.unidad_id = unidad.id
            LEFT JOIN celular ON equipo.celular_id = celular.id
            WHERE equipo.id = ?
        """, (equipo_id,))
        equipo_detalle = cursor.fetchone()

        if not equipo_detalle:
            flash('No se encontró el equipo', 'error')
            return redirect(url_for('index'))

        cursor.execute("""
            SELECT TOP 3
                asignacion_equipo.fecha_asignacion,
                asignacion_equipo.fecha_devolucion,
                asignacion_equipo.observaciones,
                asignacion_equipo.archivo_pdf,
                asignacion_equipo.archivo_pdf_devolucion,
                persona.id AS persona_id,
                persona.nombres AS persona_nombre,
                persona.apellidos AS persona_apellidos
            FROM asignacion_equipo
            JOIN persona ON asignacion_equipo.persona_id = persona.id
            WHERE asignacion_equipo.equipo_id = ?
            ORDER BY asignacion_equipo.fecha_asignacion DESC
        """, (equipo_id,))
        historial = cursor.fetchall()

        persona_id = None
        if historial:
            persona_id = historial[0].persona_id

        otros_equipos = []
        if persona_id:
            cursor.execute("""
                SELECT
                    equipo.id,
                    equipo.fcreacion,
                    equipo.observaciones,
                    estadoequipo.nombre AS estado_nombre,
                    COALESCE(unidad.nombre_e, celular.nombre) AS nombre_equipo
                FROM asignacion_equipo
                JOIN equipo ON asignacion_equipo.equipo_id = equipo.id
                JOIN estadoequipo ON equipo.estadoequipo_id = estadoequipo.id
                LEFT JOIN unidad ON equipo.unidad_id = unidad.id
                LEFT JOIN celular ON equipo.celular_id = celular.id
                WHERE asignacion_equipo.persona_id = ? AND equipo.id != ?
            """, (persona_id, equipo_id))
            otros_equipos = cursor.fetchall()

        cursor.close()

        return render_template('detalle_equipo.html', equipo=[equipo_detalle], historial=historial, otros_equipos=otros_equipos)
    
    except pyodbc.Error as ex:
        print('Error al obtener los detalles del equipo: {ex}')
        flash('Error al obtener los detalles del equipo', 'error')
        return redirect(url_for('index'))


#-----------------------------------------------------------------------
@app.route('/equipos')
@login_required
def lista_equipos():
    buscar = request.args.get('buscar', '')
    tipo_equipo = request.args.get('tipo_equipo', '')
    estado = request.args.get('estado', '')

    try:
        cursor = conexion.cursor()

        query = """
            SELECT
                equipo.id,
                tipoequipo.nombre AS tipo_equipo_nombre,
                COALESCE(unidad.nombre_e, celular.nombre) AS nombre_equipo,
                estadoequipo.nombre AS estado_nombre,
                equipo.observaciones
            FROM equipo
            LEFT JOIN tipoequipo ON equipo.tipoequipo_id = tipoequipo.id
            LEFT JOIN estadoequipo ON equipo.estadoequipo_id = estadoequipo.id
            LEFT JOIN unidad ON equipo.unidad_id = unidad.id
            LEFT JOIN celular ON equipo.celular_id = celular.id
            WHERE 1=1
        """
        
        params = []
        
        if buscar:
            query += " AND (unidad.nombre_e LIKE ? OR celular.nombre LIKE ?)"
            params.append(f'%{buscar}%')
            params.append(f'%{buscar}%')
        
        if tipo_equipo:
            query += " AND equipo.tipoequipo_id = ?"
            params.append(tipo_equipo)
        
        if estado:
            query += " AND equipo.estadoequipo_id = ?"
            params.append(estado)
        
        cursor.execute(query, params)
        equipos = cursor.fetchall()
        
        cursor.execute("SELECT id, nombre FROM tipoequipo")
        tipos_equipo = cursor.fetchall()
        
        cursor.execute("SELECT id, nombre FROM estadoequipo")
        estados = cursor.fetchall()
        
        cursor.close()

        return render_template('lista_equipos.html', equipos=equipos, tipos_equipo=tipos_equipo, estados=estados)
    
    except pyodbc.Error as ex:
        print('Error al obtener la lista de equipos: {ex}')
        flash('Error al obtener la lista de equipos', 'error')
        return redirect(url_for('index'))

#-----------------------------------------------------------------------

def allowed_file(filename):
    return '.' in filename and \
          filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/descargar_pdf/<pdf_filename>')
def descargar_pdf(pdf_filename):
    try:
        return send_from_directory(app.config['UPLOAD_FOLDER'], pdf_filename, as_attachment=True)
    except FileNotFoundError:
        flash('El archivo PDF no se encontró', 'error')
        return redirect(url_for('index'))
#-----------------------------------------

@app.route('/static/<path:filename>')
def custom_static(filename):
    return send_from_directory(app.static_folder, filename)
#------------------------------------------


def fetch_data(equipo_id):
    cursor = conexion.cursor()

    cursor.execute("""
        SELECT
            equipo.id,
            estadoequipo.nombre AS estado_nombre,
            equipo.fcreacion AS fecha_creacion,
            equipo.observaciones,
            equipo.unidad_id,
            unidad.nombre_e,
            unidad.marca,
            unidad.modelo,
            unidad.ram,
            unidad.procesador,
            unidad.almc,
            unidad.perif,
            unidad.numsello,
            unidad.serial,
            unidad.numproducto,
            unidad.tipoimpresion,
            unidad.cantidad,
            equipo.celular_id,
            celular.nombre AS nombre_celular,
            celular.marca AS marca_celular,
            celular.modelo AS modelo_celular,
            celular.serial AS serial_celular,
            celular.imei1,
            celular.imei2,
            celular.ntelefono,
            persona.nombres AS persona_nombre,
            persona.apellidos AS persona_apellidos,
            persona.area AS persona_area
        FROM equipo
        LEFT JOIN estadoequipo ON equipo.estadoequipo_id = estadoequipo.id
        LEFT JOIN unidad ON equipo.unidad_id = unidad.id
        LEFT JOIN celular ON equipo.celular_id = celular.id
        LEFT JOIN asignacion_equipo ON equipo.id = asignacion_equipo.equipo_id
        LEFT JOIN persona ON asignacion_equipo.persona_id = persona.id
        WHERE equipo.id = ?
    """, (equipo_id,))
    equipo_detalle = cursor.fetchone()
    cursor.close()
    conexion.close()

    return equipo_detalle




#-------------------------------------
@app.route('/exportar_equipos_pdf', methods=['GET'])
@login_required
def exportar_equipos_pdf():
    try:
        buscar = request.args.get('buscar', '')
        tipo_equipo = request.args.get('tipo_equipo', '')
        estado = request.args.get('estado', '')

        # Crear la consulta SQL con los filtros aplicados
        query = """
            SELECT
                equipo.id,
                tipoequipo.nombre AS tipo_equipo_nombre,
                COALESCE(unidad.nombre_e, celular.nombre) AS nombre_equipo,
                estadoequipo.nombre AS estado_nombre,
                equipo.observaciones
            FROM equipo
            LEFT JOIN tipoequipo ON equipo.tipoequipo_id = tipoequipo.id
            LEFT JOIN estadoequipo ON equipo.estadoequipo_id = estadoequipo.id
            LEFT JOIN unidad ON equipo.unidad_id = unidad.id
            LEFT JOIN celular ON equipo.celular_id = celular.id
            WHERE 1=1
        """
        
        params = []
        
        if buscar:
            query += " AND (unidad.nombre_e LIKE ? OR celular.nombre LIKE ?)"
            params.append(f'%{buscar}%')
            params.append(f'%{buscar}%')
        
        if tipo_equipo:
            query += " AND equipo.tipoequipo_id = ?"
            params.append(tipo_equipo)
        
        if estado:
            query += " AND equipo.estadoequipo_id = ?"
            params.append(estado)
        
        cursor = conexion.cursor()
        cursor.execute(query, params)
        equipos = cursor.fetchall()
        cursor.close()

        # Crear un buffer de memoria para almacenar el PDF
        buffer = io.BytesIO()

        # Crear un documento PDF
        pdf = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        # Datos para la tabla PDF
        data = [['ID', 'Tipo de Equipo', 'Nombre', 'Estado', 'Observaciones']]

        # Recorrer las filas seleccionadas y agregar los datos a la tabla
        for equipo in equipos:
            data.append([equipo.id, equipo.tipo_equipo_nombre, equipo.nombre_equipo, equipo.estado_nombre, equipo.observaciones])

        # Crear la tabla PDF
        table = Table(data)

        # Estilo de la tabla
        style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)])

        table.setStyle(style)
        elements.append(table)

        # Construir el PDF
        pdf.build(elements)

        # Devolver el PDF como una respuesta
        buffer.seek(0)
        response = make_response(buffer.getvalue())
        response.headers['Content-Disposition'] = 'attachment; filename=equipos.pdf'
        response.headers['Content-Type'] = 'application/pdf'

        return response
    except Exception as ex:
        print('Error al exportar a PDF:', ex)
        flash('Error al exportar a PDF', 'error')
        return redirect(url_for('lista_equipos'))


#---------------------------------------

#---------------------------------------
def validar_credenciales(username, password):
    try:
        cursor = conexion.cursor()
        cursor.execute('SELECT * FROM Usuarios WHERE username = ? AND password = ?', (username, password))
        usuario = cursor.fetchone()
        cursor.close()
        return usuario is not None
    except Exception as e:
        print("Error de base de datos:", e)
        return False
#----------------------------------------

#----------------------------------------
@app.route('/exportar_excel', methods=['GET'])
def exportar_excel():
    try:
        search_term = request.args.get('search_term', '')
        
        # Crear la consulta SQL con el término de búsqueda
        query = '''
            SELECT DISTINCT 
                persona.id, 
                persona.nombres, 
                persona.apellidos, 
                persona.correo, 
                persona.rut, 
                persona.dv, 
                area.nombre AS area_nombre,
                tipoequipo.nombre AS tipo_nombre,
                ISNULL(unidad.nombre_e, '') AS nombre_unidad, 
                ISNULL(celular.nombre, '') AS nombre_celular, 
                equipo.id AS equipo_id
            FROM 
                persona
            INNER JOIN 
                area ON persona.area_id = area.id
            LEFT JOIN 
                asignacion_equipo ON persona.id = asignacion_equipo.persona_id
            LEFT JOIN 
                equipo ON asignacion_equipo.equipo_id = equipo.id
            LEFT JOIN 
                unidad ON equipo.unidad_id = unidad.id
            LEFT JOIN 
                celular ON equipo.celular_id = celular.id
            LEFT JOIN 
                tipoequipo ON equipo.tipoequipo_id = tipoequipo.id
            WHERE
                equipo.estadoequipo_id = 2 AND asignacion_equipo.fecha_devolucion IS NULL
                AND (
                    persona.nombres LIKE ? OR
                    persona.apellidos LIKE ? OR
                    persona.correo LIKE ? OR
                    persona.rut LIKE ? OR
                    area.nombre LIKE ? OR
                    tipoequipo.nombre LIKE ? OR
                    unidad.nombre_e LIKE ? OR
                    celular.nombre LIKE ?
                )
        '''

        cursor = conexion.cursor()
        cursor.execute(query, ('%' + search_term + '%',) * 8)
        rows = cursor.fetchall()
        cursor.close()

        personas_dict = {}

        for row in rows:
            persona_id = row.id
            if persona_id not in personas_dict:
                personas_dict[persona_id] = {
                    'id': row.id,
                    'nombres': row.nombres,
                    'apellidos': row.apellidos,
                    'correo': row.correo,
                    'rut': row.rut,
                    'dv': row.dv,
                    'area_nombre': row.area_nombre,
                    'equipos_asignados': []
                }
            equipo = {
                'nombre': row.nombre_unidad if row.nombre_unidad else row.nombre_celular,
                'tipoequipo': row.tipo_nombre
            }
            if equipo not in personas_dict[persona_id]['equipos_asignados']:
                personas_dict[persona_id]['equipos_asignados'].append(equipo)

        wb = Workbook()
        ws = wb.active

        ws.append(['Rut', 'Nombre', 'Apellido', 'Correo', 'Área', 'Equipos Asignados'])

        for persona in personas_dict.values():
            equipos_asignados = ', '.join([f"{equipo['nombre']} - {equipo['tipoequipo']}" for equipo in persona['equipos_asignados']])
            ws.append([persona['rut'], persona['nombres'], persona['apellidos'], persona['correo'], persona['area_nombre'], equipos_asignados])

        for col in ws.columns:
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col[0].column_letter].width = adjusted_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, download_name="index.xlsx", as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as ex:
        print('Error al exportar a Excel:', ex)
        flash('Error al exportar a Excel', 'error')
        return redirect(url_for('index'))

#----
@app.route('/exportar_equipos_excel', methods=['GET'])
@login_required
def exportar_equipos_excel():
    try:
        buscar = request.args.get('buscar', '')
        tipo_equipo = request.args.get('tipo_equipo', '')
        estado = request.args.get('estado', '')

        # Crear la consulta SQL con los parámetros de búsqueda
        query = """
            SELECT
                equipo.id,
                tipoequipo.nombre AS tipo_equipo_nombre,
                COALESCE(unidad.nombre_e, celular.nombre) AS nombre_equipo,
                estadoequipo.nombre AS estado_nombre,
                equipo.observaciones
            FROM equipo
            LEFT JOIN tipoequipo ON equipo.tipoequipo_id = tipoequipo.id
            LEFT JOIN estadoequipo ON equipo.estadoequipo_id = estadoequipo.id
            LEFT JOIN unidad ON equipo.unidad_id = unidad.id
            LEFT JOIN celular ON equipo.celular_id = celular.id
            WHERE 1=1
        """

        params = []

        if buscar:
            query += " AND (unidad.nombre_e LIKE ? OR celular.nombre LIKE ?)"
            params.append(f'%{buscar}%')
            params.append(f'%{buscar}%')

        if tipo_equipo:
            query += " AND equipo.tipoequipo_id = ?"
            params.append(tipo_equipo)

        if estado:
            query += " AND equipo.estadoequipo_id = ?"
            params.append(estado)

        cursor = conexion.cursor()
        cursor.execute(query, params)
        equipos = cursor.fetchall()
        cursor.close()

        # Crear un libro de Excel y seleccionar la primera hoja
        wb = Workbook()
        ws = wb.active

        # Añadir encabezados a la primera fila
        ws.append(['ID', 'Tipo de Equipo', 'Nombre', 'Estado', 'Observaciones'])

        # Añadir filas de datos desde la consulta SQL
        for equipo in equipos:
            ws.append([equipo.id, equipo.tipo_equipo_nombre, equipo.nombre_equipo, equipo.estado_nombre, equipo.observaciones])

        # Ajustar el ancho de las columnas
        for col in ws.columns:
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col[0].column_letter].width = adjusted_width

        # Crear una respuesta con el archivo Excel adjunto
        filename = "equipos.xlsx"
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = 'attachment; filename=equipos.xlsx'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        return response
    except Exception as ex:
        print('Error al exportar a Excel:', ex)
        flash('Error al exportar a Excel', 'error')
        return redirect(url_for('lista_equipos'))





def obtener_personas():
    cursor = conexion.cursor()
    try:
        search_term = ''  # Puedes ajustar este valor según sea necesario

        query = '''
            SELECT DISTINCT 
                persona.id, 
                persona.nombres, 
                persona.apellidos, 
                persona.correo, 
                persona.rut, 
                persona.dv, 
                area.nombre AS area_nombre,
                tipoequipo.nombre AS tipo_nombre,
                ISNULL(unidad.nombre_e, '') AS nombre_unidad, 
                ISNULL(celular.nombre, '') AS nombre_celular, 
                equipo.id AS equipo_id
            FROM 
                persona
            INNER JOIN 
                area ON persona.area_id = area.id
            LEFT JOIN 
                asignacion_equipo ON persona.id = asignacion_equipo.persona_id
            LEFT JOIN 
                equipo ON asignacion_equipo.equipo_id = equipo.id
            LEFT JOIN 
                unidad ON equipo.unidad_id = unidad.id
            LEFT JOIN 
                celular ON equipo.celular_id = celular.id
            LEFT JOIN 
                tipoequipo ON equipo.tipoequipo_id = tipoequipo.id
            WHERE
                equipo.estadoequipo_id = 2 AND asignacion_equipo.fecha_devolucion IS NULL
                AND (
                    persona.nombres LIKE ? OR
                    persona.apellidos LIKE ? OR
                    persona.correo LIKE ? OR
                    persona.rut LIKE ? OR
                    area.nombre LIKE ? OR
                    tipoequipo.nombre LIKE ? OR
                    unidad.nombre_e LIKE ? OR
                    celular.nombre LIKE ?
                )
        '''

        cursor.execute(query, ('%' + search_term + '%',) * 8)  # Aplicar el término de búsqueda a cada campo

        rows = cursor.fetchall()
        cursor.close()

        personas = {}
        for row in rows:
            persona_id = row.id
            if persona_id not in personas:
                personas[persona_id] = {
                    'id': row.id,
                    'nombres': row.nombres,
                    'apellidos': row.apellidos,
                    'correo': row.correo,
                    'rut': row.rut,
                    'dv': row.dv,
                    'area_nombre': row.area_nombre,
                    'equipos_asignados': []
                }
            equipo = {
                'equipo_id': row.equipo_id,
                'nombre': row.nombre_unidad if row.nombre_unidad else row.nombre_celular,
                'tipoequipo': row.tipo_nombre
            }
            if equipo['nombre']:
                personas[persona_id]['equipos_asignados'].append(equipo)

        return list(personas.values())
    except pyodbc.Error as ex:
        print('Error al obtener los datos:', ex)
        flash('Error al cargar la página de asignación', 'error')
        return []

@app.route('/exportar_reporte_individual/<int:persona_id>', methods=['GET'])
@login_required
def exportar_reporte_individual(persona_id):
    try:
        # Obtener los datos de la persona por su ID
        persona = obtener_persona_por_id(persona_id)
        
        if persona:
            # Crear un nuevo libro de Excel y seleccionar la primera hoja
            wb = Workbook()
            ws = wb.active

            # Añadir encabezados a la primera fila
            ws.append(['Rut', 'Nombre', 'Apellido', 'Correo', 'Área', 'Equipos Asignados'])

            # Concatenar los nombres de los equipos asignados en una sola cadena
            equipos_asignados = ', '.join([f"{equipo['nombre']} - {equipo['tipoequipo']}" for equipo in persona['equipos_asignados']])

            # Agregar los datos de la persona y los equipos asignados en una fila
            ws.append([persona['rut'], persona['nombres'], persona['apellidos'], persona['correo'], persona['area_nombre'], equipos_asignados])

            # Ajustar el ancho de las columnas
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[col[0].column_letter].width = adjusted_width

            # Crear una respuesta con el archivo Excel adjunto
            filename = f"reporte_{persona_id}.xlsx"
            wb.save(filename)
            with open(filename, 'rb') as file:
                response = make_response(file.read())
            response.headers['Content-Disposition'] = f'attachment; filename=reporte_{persona_id}.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

            return response
        else:
            flash('Persona no encontrada', 'error')
            return redirect(url_for('index'))
    except Exception as ex:
        print('Error al exportar el reporte individual a Excel:', ex)
        flash('Error al exportar el reporte individual a Excel', 'error')
        return redirect(url_for('index'))


#---------------------------------------------

def obtener_persona_por_id(persona_id):
    cursor = conexion.cursor()
    try:
        query = '''
            SELECT DISTINCT 
                persona.id, 
                persona.nombres, 
                persona.apellidos, 
                persona.correo, 
                persona.rut, 
                persona.dv, 
                area.nombre AS area_nombre
            FROM 
                persona
            INNER JOIN 
                area ON persona.area_id = area.id
            WHERE
                persona.id = ?
        '''

        cursor.execute(query, (persona_id,))
        row = cursor.fetchone()

        if row:
            persona = {
                'id': row.id,
                'nombres': row.nombres,
                'apellidos': row.apellidos,
                'correo': row.correo,
                'rut': row.rut,
                'dv': row.dv,
                'area_nombre': row.area_nombre,
                'equipos_asignados': obtener_equipos_por_persona(persona_id)
            }
            return persona
        else:
            return None
    except pyodbc.Error as ex:
        print('Error al obtener la persona por ID:', ex)
        return None
    finally:
        cursor.close()

def obtener_equipos_por_persona(persona_id):
    cursor = conexion.cursor()
    try:
        query = '''
            SELECT 
                equipo.id AS equipo_id,
                ISNULL(unidad.nombre_e, '') AS nombre_unidad, 
                ISNULL(celular.nombre, '') AS nombre_celular,
                tipoequipo.nombre AS tipo_nombre
            FROM 
                asignacion_equipo
            INNER JOIN 
                equipo ON asignacion_equipo.equipo_id = equipo.id
            LEFT JOIN 
                unidad ON equipo.unidad_id = unidad.id
            LEFT JOIN 
                celular ON equipo.celular_id = celular.id
            LEFT JOIN 
                tipoequipo ON equipo.tipoequipo_id = tipoequipo.id
            WHERE 
                asignacion_equipo.persona_id = ?
        '''

        cursor.execute(query, (persona_id,))
        rows = cursor.fetchall()

        equipos_asignados = []
        for row in rows:
            equipo = {
                'equipo_id': row.equipo_id,
                'nombre': row.nombre_unidad if row.nombre_unidad else row.nombre_celular,
                'tipoequipo': row.tipo_nombre
            }
            equipos_asignados.append(equipo)

        return equipos_asignados
    except pyodbc.Error as ex:
        print('Error al obtener los equipos por persona:', ex)
        return []
    finally:
        cursor.close()

##############################

@app.route('/exportar_pdf', methods=['GET'])
def exportar_pdf():
    try:
        search_term = request.args.get('search_term', '')

        # Crear la consulta SQL con el término de búsqueda
        query = '''
            SELECT DISTINCT 
                persona.id, 
                persona.nombres, 
                persona.apellidos, 
                persona.correo, 
                area.nombre AS area_nombre,
                tipoequipo.nombre AS tipo_nombre,
                ISNULL(unidad.nombre_e, '') AS nombre_unidad, 
                ISNULL(celular.nombre, '') AS nombre_celular, 
                equipo.id AS equipo_id
            FROM 
                persona
            INNER JOIN 
                area ON persona.area_id = area.id
            LEFT JOIN 
                asignacion_equipo ON persona.id = asignacion_equipo.persona_id
            LEFT JOIN 
                equipo ON asignacion_equipo.equipo_id = equipo.id
            LEFT JOIN 
                unidad ON equipo.unidad_id = unidad.id
            LEFT JOIN 
                celular ON equipo.celular_id = celular.id
            LEFT JOIN 
                tipoequipo ON equipo.tipoequipo_id = tipoequipo.id
            WHERE
                equipo.estadoequipo_id = 2 AND asignacion_equipo.fecha_devolucion IS NULL
                AND (
                    persona.nombres LIKE ? OR
                    persona.apellidos LIKE ? OR
                    persona.correo LIKE ? OR
                    
                    area.nombre LIKE ? OR
                    tipoequipo.nombre LIKE ? OR
                    unidad.nombre_e LIKE ? OR
                    celular.nombre LIKE ?
                )
        '''

        cursor = conexion.cursor()
        cursor.execute(query, ('%' + search_term + '%',) * 7)  # Aplicar el término de búsqueda a cada campo
        rows = cursor.fetchall()
        cursor.close()

        # Crear un buffer de memoria para almacenar el PDF
        buffer = io.BytesIO()

        # Crear un documento PDF
        pdf = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        # Datos para la tabla PDF
        data = [['Nombre', 'Apellido', 'Correo', 'Área', 'Equipos Asignados']]

        # Recorrer las filas seleccionadas y agregar los datos a la tabla
        for row in rows:
            equipos_asignados = ', '.join([f"{row.nombre_unidad if row.nombre_unidad else row.nombre_celular} - {row.tipo_nombre}"])
            data.append([row.nombres, row.apellidos, row.correo, row.area_nombre, equipos_asignados])

        # Crear la tabla PDF
        table = Table(data)

        # Estilo de la tabla
        style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)])

        table.setStyle(style)
        elements.append(table)

        # Construir el PDF
        pdf.build(elements)

        # Devolver el PDF como una respuesta
        buffer.seek(0)
        response = make_response(buffer.getvalue())
        response.headers['Content-Disposition'] = 'attachment; filename=exportacion.pdf'
        response.headers['Content-Type'] = 'application/pdf'

        return response
    except Exception as ex:
        print('Error al exportar a PDF:', ex)
        flash('Error al exportar a PDF', 'error')
        return redirect(url_for('index'))


class PDF(FPDF):
    def header(self):

        left_margin = 20
        self.set_left_margin(left_margin)
        self.set_x(left_margin)
        base_dir = os.path.abspath(os.path.dirname(__file__))
        imagen_path = os.path.join(base_dir, 'static', 'logochf.png')
        self.image(imagen_path, x=140, y=12, w=50)        
        self.set_font('Helvetica', 'B', 25)
        self.cell(0, 40, 'Carta de Devolucion', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')

    def footer(self):
        self.set_y(-50)
    
        # Ancho total de la página
        page_width = self.w - 2 * self.l_margin

        # Ancho de cada celda de firma
        cell_width = page_width / 3


        self.cell(cell_width, 8, '_____________', 0, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C')
        self.cell(cell_width, 8, '_____________', 0, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C')
        self.cell(cell_width, 8, '_____________', 0, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
        self.set_font('Helvetica', 'B', 12)
        self.cell(cell_width, 5, 'FIRMA DEL USUARIO', 0, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C')
        self.cell(cell_width, 5, 'GASTON MARILEO', 0, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C')
        self.cell(cell_width, 5, 'CARLOS AHUMADA', 0, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
        self.set_font('Helvetica', 'B', 8)
        self.cell(cell_width, 5, 'RESPONSABLE', 0, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C')
        self.cell(cell_width, 5, 'JEFE SISTEMA TI', 0, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C')
        self.cell(cell_width, 5, 'ENCARGADO SOPORTE TI', 0, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')

        # Línea de firma
        

@app.route('/exportar_pdf/<int:persona_id>/<int:equipo_id>', methods=['GET'])
def exportar_pdf_persona(persona_id, equipo_id):
    try:
        query = '''
            SELECT DISTINCT 
                persona.id, 
                persona.nombres, 
                persona.apellidos, 
                persona.correo, 
                persona.rut, 
                persona.dv, 
                area.nombre AS area_nombre,
                tipoequipo.nombre AS tipo_nombre,
                ISNULL(unidad.nombre_e, '') AS nombre_unidad, 
                ISNULL(celular.nombre, '') AS nombre_celular, 
                equipo.id AS equipo_id,
                unidad.id AS unidad_id, 
                unidad.marca AS marca,
                unidad.modelo AS modelo,
                unidad.ram AS ram,
                unidad.procesador AS procesador,
                unidad.almc AS almc,
                unidad.perif AS perif,
                unidad.numsello AS numsello,
                unidad.serial AS serial,
                unidad.numproducto AS numproducto,
                unidad.tipoimpresion AS tipoimpresion,
                unidad.cantidad AS cantidad,
                celular.id AS celular_id,
                celular.marca AS marca_celular,
                celular.modelo AS modelo_celular,
                celular.serial AS serial_celular,
                celular.imei1 AS imei1,
                celular.imei2 AS imei2,
                celular.ntelefono AS ntelefono
            FROM 
                persona
            INNER JOIN 
                area ON persona.area_id = area.id
            LEFT JOIN 
                asignacion_equipo ON persona.id = asignacion_equipo.persona_id
            LEFT JOIN 
                equipo ON asignacion_equipo.equipo_id = equipo.id
            LEFT JOIN 
                unidad ON equipo.unidad_id = unidad.id
            LEFT JOIN 
                celular ON equipo.celular_id = celular.id
            LEFT JOIN 
                tipoequipo ON equipo.tipoequipo_id = tipoequipo.id
            WHERE
                persona.id = ? AND equipo.id = ?
        '''

        cursor = conexion.cursor()
        cursor.execute(query, (persona_id, equipo_id))
        rows = cursor.fetchall()
        cursor.close()

        if not rows:
            flash('No se encontraron datos para el equipo seleccionado', 'error')
            return redirect(url_for('index'))

        buffer = io.BytesIO()

        # Crear un documento PDF
        pdf = PDF()
        pdf.add_page()

        # Agregar borde cuadrado
        pdf.set_line_width(0.5)
        pdf.rect(10, 10, 190, 270)

        left_margin = 20

        nombre = f"{rows[0].nombres} {rows[0].apellidos}"
        area = rows[0].area_nombre
        pdf.set_font('Helvetica', 'B', 16)
        pdf.set_x(left_margin)
        pdf.cell(0, 10, f'Se recibe de Sr/a: {nombre}', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')
        pdf.set_x(left_margin)
        pdf.set_font('Helvetica', '', 14)
        pdf.cell(0, 2, f'{area}', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')

        # Espacio
        pdf.cell(0, 10, '', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.set_font('Helvetica', '', 16)

        # Agregar datos del equipo
        for row in rows:
            equipo_nombre = row.nombre_unidad if row.nombre_unidad else row.nombre_celular
            tipo_nombre = row.tipo_nombre
            if equipo_nombre:
                pdf.cell(0, 6, f"Nombre del equipo: {equipo_nombre}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')
            if tipo_nombre:
                pdf.cell(0, 6, f"Tipo de equipo: {tipo_nombre}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')

            # Agregar otros atributos del equipo si están presentes
            if row.unidad_id:
                if row.marca:
                    pdf.cell(0, 6, f"Marca {row.marca}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')
                if row.modelo:
                    pdf.cell(0, 6, f"Modelo {row.modelo}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')
                if row.ram:
                    pdf.cell(0, 6, f"RAM {row.ram}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')
                if row.procesador:
                    pdf.cell(0, 6, f"Procesador {row.procesador}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')
                if row.almc:
                    pdf.cell(0, 6, f"Almacenamiento {row.almc}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')
                if row.perif:
                    pdf.cell(0, 6, f"Periféricos {row.perif}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')
                if row.numsello:
                    pdf.cell(0, 6, f"N° Sello {row.numsello}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')
                if row.serial:
                    pdf.cell(0, 6, f"S/N {row.serial}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')
                if row.numproducto:
                    pdf.cell(0, 6, f"N° de Producto {row.numproducto}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')
                if row.tipoimpresion:
                    pdf.cell(0, 6, f"Tipo de Impresión {row.tipoimpresion}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')
                if row.cantidad:
                    pdf.cell(0, 6, f"Cantidad {row.cantidad}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')
            if row.celular_id:
                if row.marca_celular:
                    pdf.cell(0, 6, f"Marca {row.marca_celular}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')
                if row.modelo_celular:
                    pdf.cell(0, 6, f"Modelo {row.modelo_celular}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')
                if row.serial_celular:
                    pdf.cell(0, 6, f"Serial {row.serial_celular}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')
                if row.imei1:
                    pdf.cell(0, 6, f"IMEI 1 {row.imei1}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')
                if row.imei2:
                    pdf.cell(0, 6, f"IMEI 2 {row.imei2}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')
                if row.ntelefono:
                    pdf.cell(0, 6, f"Linea {row.ntelefono}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')
            pdf.cell(0, 8, '', new_x=XPos.LMARGIN, new_y=YPos.NEXT)

        # Guardar el PDF en el buffer
        pdf.output(buffer)

        nombre_archivo = f'carta_devolucion_{rows[0].nombres}_{rows[0].apellidos}.pdf'.replace(' ', '_')


        # Devolver el PDF como una respuesta
        buffer.seek(0)
        response = make_response(buffer.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename={nombre_archivo}'
        response.headers['Content-Type'] = 'application/pdf'

        return response
    except Exception as ex:
        print(f'Error al exportar a PDF: {ex}')
        flash('Error al exportar a PDF', 'error')
        return redirect(url_for('index'))
#-------------------------------------------------------------------------------

@app.route('/exportar_pdf_varios/<int:persona_id>', methods=['GET'])
def exportar_pdf_varios(persona_id):
    try:
        query = '''
            SELECT DISTINCT 
                persona.id, 
                persona.nombres, 
                persona.apellidos, 
                persona.correo, 
                persona.rut, 
                persona.dv, 
                area.nombre AS area_nombre,
                tipoequipo.nombre AS tipo_nombre,
                ISNULL(unidad.nombre_e, '') AS nombre_unidad, 
                ISNULL(celular.nombre, '') AS nombre_celular, 
                equipo.id AS equipo_id,
                unidad.id AS unidad_id, 
                unidad.marca AS marca,
                unidad.modelo AS modelo,
                unidad.ram AS ram,
                unidad.procesador AS procesador,
                unidad.almc AS almc,
                unidad.perif AS perif,
                unidad.numsello AS numsello,
                unidad.serial AS serial,
                unidad.numproducto AS numproducto,
                unidad.tipoimpresion AS tipoimpresion,
                unidad.cantidad AS cantidad,
                celular.id AS celular_id,
                celular.marca AS marca_celular,
                celular.modelo AS modelo_celular,
                celular.serial AS serial_celular,
                celular.imei1 AS imei1,
                celular.imei2 AS imei2,
                celular.ntelefono AS ntelefono
            FROM 
                persona
            INNER JOIN 
                area ON persona.area_id = area.id
            LEFT JOIN 
                asignacion_equipo ON persona.id = asignacion_equipo.persona_id
            LEFT JOIN 
                equipo ON asignacion_equipo.equipo_id = equipo.id
            LEFT JOIN 
                unidad ON equipo.unidad_id = unidad.id
            LEFT JOIN 
                celular ON equipo.celular_id = celular.id
            LEFT JOIN 
                tipoequipo ON equipo.tipoequipo_id = tipoequipo.id
            WHERE
                persona.id = ?
        '''

        cursor = conexion.cursor()
        cursor.execute(query, (persona_id,))
        rows = cursor.fetchall()
        cursor.close()

        if not rows:
            flash('No se encontraron datos para los equipos asignados', 'error')
            return redirect(url_for('index'))

        # Crear un documento PDF
        pdf = PDF()
        pdf.set_auto_page_break(auto=True, margin=50)

        # Definir una función para agregar contenido a cada página
        def agregar_contenido_pagina(pdf, rows):
            pdf.add_page()
            pdf.set_line_width(0.5)
            pdf.rect(10, 10, 190, 270)

            left_margin = 20

            nombre = f"{rows[0].nombres} {rows[0].apellidos}"
            area = rows[0].area_nombre
            pdf.set_font('Helvetica', 'B', 16)
            pdf.set_x(left_margin)
            pdf.cell(0, 10, f'Se recibe de Sr/a: {nombre}', 0, 1, align='L')
            pdf.set_x(left_margin)
            pdf.set_font('Helvetica', '', 14)
            pdf.cell(0, 10, f'{area}', 0, 1, align='L')

            # Espacio
            pdf.cell(0, 10, '', 0, 1)
            pdf.set_font('Helvetica', '', 16)

            # Agregar datos de los equipos
            for row in rows:
                equipo_nombre = row.nombre_unidad if row.nombre_unidad else row.nombre_celular
                tipo_nombre = row.tipo_nombre
                if equipo_nombre:
                    pdf.cell(0, 6, f"Nombre del equipo: {equipo_nombre}", 0, 1, align='L')
                if tipo_nombre:
                    pdf.cell(0, 6, f"Tipo de equipo: {tipo_nombre}", 0, 1, align='L')

                # Agregar otros atributos del equipo si están presentes
                if row.unidad_id:
                    if row.marca:
                        pdf.cell(0, 6, f"Marca: {row.marca}", 0, 1, align='L')
                    if row.modelo:
                        pdf.cell(0, 6, f"Modelo: {row.modelo}", 0, 1, align='L')
                    if row.ram:
                        pdf.cell(0, 6, f"RAM: {row.ram}", 0, 1, align='L')
                    if row.procesador:
                        pdf.cell(0, 6, f"Procesador: {row.procesador}", 0, 1, align='L')
                    if row.almc:
                        pdf.cell(0, 6, f"Almacenamiento: {row.almc}", 0, 1, align='L')
                    if row.perif:
                        pdf.cell(0, 6, f"Periféricos: {row.perif}", 0, 1, align='L')
                    if row.numsello:
                        pdf.cell(0, 6, f"N° Sello: {row.numsello}", 0, 1, align='L')
                    if row.serial:
                        pdf.cell(0, 6, f"S/N: {row.serial}", 0, 1, align='L')
                    if row.numproducto:
                        pdf.cell(0, 6, f"N° de Producto: {row.numproducto}", 0, 1, align='L')
                    if row.tipoimpresion:
                        pdf.cell(0, 6, f"Tipo de Impresión: {row.tipoimpresion}", 0, 1, align='L')
                    if row.cantidad:
                        pdf.cell(0, 6, f"Cantidad: {row.cantidad}", 0, 1, align='L')
                if row.celular_id:
                    if row.marca_celular:
                        pdf.cell(0, 6, f"Marca: {row.marca_celular}", 0, 1, align='L')
                    if row.modelo_celular:
                        pdf.cell(0, 6, f"Modelo: {row.modelo_celular}", 0, 1, align='L')
                    if row.serial_celular:
                        pdf.cell(0, 6, f"Serial: {row.serial_celular}", 0, 1, align='L')
                    if row.imei1:
                        pdf.cell(0, 6, f"IMEI 1: {row.imei1}", 0, 1, align='L')
                    if row.imei2:
                        pdf.cell(0, 6, f"IMEI 2: {row.imei2}", 0, 1, align='L')
                    if row.ntelefono:
                        pdf.cell(0, 6, f"Linea: {row.ntelefono}", 0, 1, align='L')
                pdf.cell(0, 8, '', 0, 1)

        # Agregar contenido a la primera página
        agregar_contenido_pagina(pdf, rows)

        # Guardar el PDF en un buffer de bytes
        pdf_output = io.BytesIO()
        pdf.output(pdf_output)
        pdf_bytes = pdf_output.getvalue()
        pdf_output.close()

        # Devolver el PDF como una respuesta
        response = make_response(pdf_bytes)
        response.headers['Content-Disposition'] = f'attachment; filename=carta_devolucion_varios_equipos.pdf'
        response.headers['Content-Type'] = 'application/pdf'

        return response

    except Exception as ex:
        print(f'Error al exportar a PDF: {ex}')
        flash('Error al exportar a PDF', 'error')
        return redirect(url_for('index'))








###############################
###############################
###############################

if __name__ == '__main__':
    app.run(host=ruta, port=5000)

